from openpyxl import Workbook, load_workbook
from pprint import pprint
wb2 = load_workbook('sample.xlsx')
ws=wb2.active
ws["B3"]=2

# cell_range = ws['A1':]
for i in tuple(ws.columns)[0]:
    pprint(i.value)
# print(wb2.sheetnames[0])
# print(tuple(ws.rows))


